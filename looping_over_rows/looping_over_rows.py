import openpyxl as xl

wb = xl.load_workbook('example.xlsx', data_only=True)
ws = wb['Sheet1']

# get the headers
headers = []
for row in ws.iter_rows(min_row=1, max_row=1):
    for cell in row:
        headers.append(cell.value)

# loop over rows 2 to 10
for row in ws.iter_rows(min_row=2, max_row=5):
    # store the row data in a dictionary
    entry = {}
    for cell in row:
        index = cell.column - 1
        entry[headers[index]] = cell.value
    # inspect the data
    print(entry)

    
    # insert logic here
    
    # modify the sheet: add a value to the next column to the right
    ws.cell(cell.row, len(row) + 1).value = "done"

wb.save('example_modified.xlsx')
wb.close()
